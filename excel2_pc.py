"""
使用openpyxl实现以下需求

使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
(可以做一部分优化)
"""
# 导入openpyxl库及使用的类
from openpyxl import Workbook

from openpyxl import load_workbook


# excel表练习类
class PracticeExcel:
    # 创建表方法
    def create_table(self):
        # 实例化excel表写入类
        wb = Workbook()
        ws1 = wb.active
        # 定义页名与表头
        ws1.title = "body_data"
        ws1["A1"] = "姓名"
        ws1["B1"] = "身高"
        ws1["C1"] = "体重"
        ws1["D1"] = "健康体重"
        ws1["E1"] = "备注"
        # 设定姓名
        name = ["李峰", "张狂", "谢邀", "上官镜"]
        # 设定身高体重数据
        body_data = {180: 66, 170: 58, 175: 62, 164: 76}
        # 从字典中取出身高信息
        height = [i for i in body_data.keys()]
        # 逐条按顺序写入姓名，身高，体重等信息
        for i in range(len(height)):
            # 从第二行开始写入，按列区别信息
            ws1.cell(row=i + 2, column=1).value = name[i]
            ws1.cell(row=i + 2, column=2).value = height[i]
            # 通过字典的键来获取值
            ws1.cell(row=i + 2, column=3).value = body_data[height[i]]
        # 保存表
        wb.save("health.xlsx")

    # 健康数据类
    def healthy_data(self):
        # 实例化读取表类，声明读取文件名称
        ld = load_workbook(filename="health.xlsx")
        # 选择读取页
        sheet = ld["body_data"]
        # 读取表，并对其中内容进行修改
        for i in range(4):
            # 读取表中内容
            name = sheet.cell(row=i + 2, column=1).value
            height = sheet.cell(row=i + 2, column=2).value
            weight = sheet.cell(row=i + 2, column=3).value
            # 健康体重计算公式
            healthy_weight = (height - 70) * 0.6
            # 健康体重判断
            if healthy_weight == weight:
                # 打印身体数据
                print("{}的体重为：{}kg".format(name, weight))
                print("健康体重为：{}kg".format(healthy_weight))
                print("体重健康\n")
                # 将健康体重与备注信息写入表中
                sheet.cell(row=i + 2, column=4, value=healthy_weight)
                sheet.cell(row=i + 2, column=5, value="体重健康")
            elif healthy_weight < weight:
                # 打印身体数据
                print("{}的体重为：{}kg".format(name, weight))
                print("健康体重为：{}kg".format(healthy_weight))
                print("体重偏重\n")
                # 将健康体重与备注信息写入表中
                sheet.cell(row=i + 2, column=4, value=healthy_weight)
                sheet.cell(row=i + 2, column=5, value="体重偏重")
            elif healthy_weight > weight:
                # 打印身体数据
                print("{}的体重为：{}kg".format(name, weight))
                print("健康体重为：{}kg".format(healthy_weight))
                print("体重偏轻\n")
                # 将健康体重与备注信息写入表中
                sheet.cell(row=i + 2, column=4, value=healthy_weight)
                sheet.cell(row=i + 2, column=5, value="体重偏轻")
        # 保存对表的修改
        ld.save(filename="health.xlsx")


# 实例化excel表练习类
pe = PracticeExcel()
# 调用其中的创建表与健康信息方法
pe.create_table()
pe.healthy_data()
